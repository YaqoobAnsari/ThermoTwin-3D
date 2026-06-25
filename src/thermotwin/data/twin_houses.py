"""Twin Houses U-value validation — our physics vs the real building's documented U.

The IEA Annex 71 Twin Houses package ships ``01_Constructions_TwinHouses.xlsx``: per envelope
element (West/East/South/North walls, ceiling, floor, roof, door, cellar wall) the full layer
build-up (thickness d [mm], conductivity λ [W/mK]) + surface films (Rsi/Rse) + the documented
U-value. That makes a clean, *quantitative*, real-data check of the conduction physics our
operator's analytic prior is built on: parse each element's layers, compute
``U = 1/(Rsi + Σ d/λ + Rse)`` with our own physics, and compare to the building's documented U.

This is the comparable metric for the Twin Houses rung — **per-element U-value MAE vs the real
documented values** (the U-readout validation; sparse/point-calibrated, not a field).
"""

from __future__ import annotations

import io
import zipfile

__all__ = ["open_constructions", "ELEMENT_SHEETS", "parse_element", "computed_u", "evaluate"]

# Envelope-element sheets carrying a layer build-up (Windows/glazing are given U directly).
ELEMENT_SHEETS = (
    "West", "East", "South", "North", "Int Walls", "ceiling", "floor", "roof", "Front door",
)


def open_constructions(path: str):
    """Load the constructions workbook from a ``.xlsx`` path (extract it first if zipped).

    The Additional-Documents zip uses a compression Python's ``zipfile`` can't decode, so the
    XLSX is extracted once with the system ``unzip`` to
    ``data/raw/twin_houses/01_Constructions_TwinHouses.xlsx``.
    """
    import openpyxl

    if str(path).endswith(".xlsx"):
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    zf = zipfile.ZipFile(path)
    member = next(n for n in zf.namelist() if n.endswith("01_Constructions_TwinHouses.xlsx"))
    return openpyxl.load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)


def parse_element(ws) -> dict:
    """Parse one element sheet -> ``{rsi, rse, layers:[(name,d_m,lambda,R)], documented_u}``."""
    rsi = rse = documented_u = None
    layers: list[tuple] = []
    in_struct = False
    started = False
    d_col, l_col, r_col = 1, 2, 3  # wall layout default; reset per table from its header
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        # First occurrence only (some sheets hold several constructions; we take the first).
        for i, c in enumerate(cells):
            if isinstance(c, str):
                s = c.strip()
                nxt = cells[i + 1] if i + 1 < len(cells) else None
                if s.startswith("Rsi") and rsi is None and isinstance(nxt, (int, float)):
                    rsi = float(nxt)
                elif s.startswith("Rse") and rse is None and isinstance(nxt, (int, float)):
                    rse = float(nxt)
                elif s == "U =" and documented_u is None and isinstance(nxt, (int, float)):
                    documented_u = float(nxt)
        c0 = str(cells[0]).strip() if cells and cells[0] is not None else ""
        if c0 == "Structure":
            if started:  # a second construction table on the same sheet -> stop at the first
                break
            # The header names the columns; some sheets (roof) omit the R column (col-3 = σ).
            hdr = [str(c).strip() if c is not None else "" for c in cells]
            d_col = hdr.index("d") if "d" in hdr else 1
            l_col = hdr.index("l") if "l" in hdr else 2
            r_col = hdr.index("R") if "R" in hdr else None
            in_struct = True
            started = True
            continue
        if in_struct:
            # Terminate the layer table only at the totals rows; the units row (empty col-0,
            # '[mm]' in col-1) and any blank row are skipped, not treated as the end.
            if "resistance" in c0.lower() or c0.startswith("Heat transfer"):
                in_struct = False
                continue
            d = cells[d_col] if d_col < len(cells) else None
            lam = cells[l_col] if l_col < len(cells) else None
            r = cells[r_col] if (r_col is not None and r_col < len(cells)) else None
            if isinstance(d, (int, float)) and isinstance(lam, (int, float)) and lam > 0:
                layers.append((c0, d / 1000.0, float(lam), (d / 1000.0) / lam))  # mm -> m
            elif isinstance(r, (int, float)) and r > 0:  # layer given by resistance (air gap, etc.)
                layers.append((c0, None, None, float(r)))
    return {"rsi": rsi, "rse": rse, "layers": layers, "documented_u": documented_u}


def computed_u(elem: dict) -> float:
    """``U = 1/(Rsi + Σ R_layer + Rse)`` — our conduction physics on the element's layers."""
    r_layers = sum(layer[3] for layer in elem["layers"])
    rsi = elem["rsi"] if elem["rsi"] is not None else 0.13
    rse = elem["rse"] if elem["rse"] is not None else 0.04
    return 1.0 / (rsi + r_layers + rse)


def evaluate(path: str) -> dict:
    """Compute our U per element, compare to the documented U; return per-element + summary."""
    wb = open_constructions(path)
    rows = []
    for name in ELEMENT_SHEETS:
        if name not in wb.sheetnames:
            continue
        elem = parse_element(wb[name])
        if not elem["layers"] or elem["documented_u"] is None:
            continue
        u_ours = computed_u(elem)
        rows.append({
            "element": name,
            "n_layers": len(elem["layers"]),
            "u_computed": round(u_ours, 4),
            "u_documented": round(elem["documented_u"], 4),
            "abs_error": round(abs(u_ours - elem["documented_u"]), 4),
        })
    import numpy as np

    err = np.array([r["abs_error"] for r in rows]) if rows else np.array([0.0])
    return {
        "elements": rows,
        "u_mae": round(float(err.mean()), 5),
        "u_max_error": round(float(err.max()), 5),
        "n_elements": len(rows),
    }
